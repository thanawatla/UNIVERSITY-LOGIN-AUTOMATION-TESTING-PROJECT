from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

EXCEL_FILE = "test_cases.xlsx"
SCREENSHOT_DIR = "screenshots"

# สร้างโฟลเดอร์ screenshots ถ้าไม่มี
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

def log_result(test_name, actual_result, page=None):
    print(f"🔹 Logging result for: {test_name}")

    # ตรวจสอบว่าไฟล์ Excel มีอยู่หรือไม่
    if not os.path.exists(EXCEL_FILE):
        print("❌ ไม่พบไฟล์ test_cases.xlsx")
        return

    # โหลดไฟล์ Excel
    workbook = load_workbook(EXCEL_FILE)
    target_sheet_name = "login01"

    # ตรวจสอบว่า login01 มีอยู่ในไฟล์หรือไม่
    if target_sheet_name not in workbook.sheetnames:
        print(f"⚠️ ไม่พบชีต '{target_sheet_name}' ในไฟล์")
        workbook.close()
        return

    sheet = workbook[target_sheet_name]

    # ตัด prefix ให้เหลือชื่อสั้น (เช่น pos01, neg02)
    short_test_name = test_name.split("_")[-1]

    found = False

    # ค้นหาชื่อ Test Case ID ในคอลัมน์ A (Test Case ID)
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        test_case_cell = row[0]
        if test_case_cell.value and (test_case_cell.value.strip() == test_name or test_case_cell.value.strip() == short_test_name):
            row_number = test_case_cell.row
            status = "PASS" if "✅" in actual_result else "FAIL"

            # อัปเดตผลลัพธ์และสถานะในคอลัมน์ G และ H
            sheet[f"G{row_number}"] = actual_result
            sheet[f"H{row_number}"] = status

            # ถ้ามีการจับภาพหน้าจอ
            if page:
                screenshot_path = os.path.join(SCREENSHOT_DIR, f"{test_name}.png")
                page.screenshot(path=screenshot_path)
                sheet[f"I{row_number}"] = screenshot_path

                # แทรกรูปภาพลงในไฟล์ Excel
                if os.path.exists(screenshot_path):
                    img = Image(screenshot_path)
                    img.width, img.height = 330, 210
                    sheet.add_image(img, f"I{row_number}")

            found = True
            print(f"✅ Updated '{test_name}' ในชีต '{target_sheet_name}' (พบจาก '{test_case_cell.value}')")
            break

    # ถ้าไม่พบ Test Case ID ในชีต
    if not found:
        print(f"⚠️ ไม่พบ Test Case ID '{test_name}' หรือ '{short_test_name}' ใน sheet '{target_sheet_name}'")

    # บันทึกไฟล์ Excel
    workbook.save(EXCEL_FILE)
    workbook.close()
    print(f"✅ Logging completed and saved in {EXCEL_FILE}")